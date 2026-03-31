# -*- coding: utf-8 -*-
"""
NBET Excel Import Wizard
Parses the legacy NBET Excel workbook and creates staging records for review
before final import into billing cycle records.

MAPPING NOTES:
- "Inputs" sheet: period-level inputs in rows (label in col A, value in col B or later)
  and GENCO operational data in a tabular section (GENCO names as row labels,
  data columns as headers). Exact layout varies by period — fuzzy header matching is used.
- "Rates" sheet: GENCO-specific rates with GENCO names as row labels.
- Unit handling: GWh values are auto-converted to kWh (x 1,000,000).
- If the workbook layout differs significantly, adjust _INPUTS_LABEL_MAP and
  _OPERATIONAL_HEADER_MAP constants below.
"""
import base64
import io
import logging
import re

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Label maps for "Inputs" sheet period-level rows ───────────────────────────
# Map of lowercased label fragments -> input_type_code
_INPUTS_LABEL_MAP = {
    'cbn central rate': 'CBN_FX_CENTRAL',
    'cbn fx central': 'CBN_FX_CENTRAL',
    'central fx': 'CBN_FX_CENTRAL',
    'cbn selling rate': 'CBN_FX_SELLING',
    'cbn fx selling': 'CBN_FX_SELLING',
    'selling rate': 'CBN_FX_SELLING',
    'tlf old': 'TLF_OLD',
    'old tlf': 'TLF_OLD',
    'tlf new': 'TLF_NEW',
    'new tlf': 'TLF_NEW',
    'hours in month': 'HOURS_IN_MONTH',
    'hours in period': 'HOURS_IN_MONTH',
    'hours': 'HOURS_IN_MONTH',
    'agip quarterly index': 'AGIP_INDEX',
    'agip index': 'AGIP_INDEX',
    'quarterly index': 'AGIP_INDEX',
    'agip exchange rate': 'AGIP_FX',
    'gas price index': 'GAS_PRICE_INDEX',
    'myto inflation': 'MYTO_INFLATION',
}

# ── Column header map for GENCO operational tabular section ───────────────────
_OPERATIONAL_HEADER_MAP = {
    'capacity sent out': 'capacity_sent_out_mw',
    'cap sent out': 'capacity_sent_out_mw',
    'sent out': 'capacity_sent_out_mw',
    'gross energy': 'gross_energy_kwh',
    'net energy': 'net_energy_kwh',
    'capacity import': 'capacity_import_mw',
    'import capacity': 'capacity_import_mw',
    'energy import': 'energy_import_kwh',
    'import energy': 'energy_import_kwh',
    'net energy import': 'net_energy_import_kwh',
    'invoiced capacity': 'invoiced_capacity_mw',
    'invoiced energy': 'invoiced_energy_kwh',
}

# ── Rates sheet column header map ─────────────────────────────────────────────
_RATES_HEADER_MAP = {
    'capacity rate': 'capacity_rate',
    'cap rate': 'capacity_rate',
    'energy rate': 'energy_rate',
    'eng rate': 'energy_rate',
    'applied fx': 'fx_rate_used',
    'fx rate': 'fx_rate_used',
    'applied tlf': 'tlf_used',
    'tlf': 'tlf_used',
    'applied index': 'index_value_used',
    'index': 'index_value_used',
}


class NbetExcelImportWizard(models.TransientModel):
    _name = 'nbet.excel.import.wizard'
    _description = 'NBET Excel Workbook Import Wizard'

    billing_cycle_id = fields.Many2one(
        'nbet.billing.cycle', string='Billing Cycle', required=True,
    )
    excel_file = fields.Binary(string='Excel Workbook', required=True, attachment=False)
    excel_filename = fields.Char(string='Filename')
    map_inputs_sheet = fields.Boolean(string='Parse Inputs Sheet', default=True)
    map_rates_sheet = fields.Boolean(string='Parse Rates Sheet', default=True)
    inputs_sheet_name = fields.Char(string='Inputs Sheet Name', default='Inputs')
    rates_sheet_name = fields.Char(string='Rates Sheet Name', default='Rates')
    batch_id = fields.Many2one('nbet.import.batch', string='Import Batch', readonly=True)
    state = fields.Selection(
        selection=[('draft', 'Upload'), ('previewed', 'Preview'), ('done', 'Done')],
        default='draft',
    )
    preview_html = fields.Html(string='Preview', readonly=True)

    def action_preview(self):
        """Parse the Excel file and create a staging batch for review."""
        self.ensure_one()
        if not HAS_OPENPYXL:
            raise UserError(
                'The openpyxl Python package is required for Excel import.\n'
                'Install it with: pip install openpyxl'
            )
        if not self.excel_file:
            raise UserError('Please upload an Excel file.')

        raw = base64.b64decode(self.excel_file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(f'Could not open Excel file: {e}')

        # Create or reset batch
        batch_name = f'Import: {self.billing_cycle_id.name} — {self.excel_filename or "workbook"}'
        batch = self.env['nbet.import.batch'].create({
            'name': batch_name,
            'billing_cycle_id': self.billing_cycle_id.id,
            'source_filename': self.excel_filename or '',
            'state': 'draft',
        })
        self.batch_id = batch.id

        # Parse sheets
        if self.map_inputs_sheet and self.inputs_sheet_name in wb.sheetnames:
            self._parse_inputs_sheet(wb[self.inputs_sheet_name], batch)
        elif self.map_inputs_sheet:
            _logger.warning('Sheet "%s" not found in workbook.', self.inputs_sheet_name)
            self._log_error(batch, self.inputs_sheet_name, 0, '',
                            'mapping_error', f'Sheet "{self.inputs_sheet_name}" not found.')

        if self.map_rates_sheet and self.rates_sheet_name in wb.sheetnames:
            self._parse_rates_sheet(wb[self.rates_sheet_name], batch)
        elif self.map_rates_sheet:
            _logger.warning('Sheet "%s" not found in workbook.', self.rates_sheet_name)

        # Build preview HTML
        self.preview_html = self._build_preview_html(batch)
        batch.state = 'previewed'
        self.state = 'previewed'

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nbet.excel.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_confirm_import(self):
        """Commit staged batch lines to actual billing records."""
        self.ensure_one()
        if not self.batch_id:
            raise UserError('No import batch found. Please run Preview first.')

        batch = self.batch_id
        cycle = self.billing_cycle_id
        imported = 0
        errors = 0

        for line in batch.line_ids:
            try:
                if line.record_type == 'cycle_input':
                    self._import_cycle_input_line(line, cycle)
                elif line.record_type == 'genco_data':
                    self._import_genco_data_line(line, cycle)
                elif line.record_type == 'rate_data':
                    self._import_rate_data_line(line, cycle)
                line.status = 'imported'
                imported += 1
            except Exception as e:
                line.status = 'error'
                line.error_message = str(e)[:250]
                errors += 1

        batch.state = 'confirmed'
        self.state = 'done'

        # Post to cycle chatter
        cycle.message_post(
            body=(
                f'Excel import completed: {imported} records imported, '
                f'{errors} errors. Batch: {batch.name}'
            )
        )

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nbet.import.batch',
            'res_id': batch.id,
            'view_mode': 'form',
        }

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet Parsers
    # ─────────────────────────────────────────────────────────────────────────

    def _parse_inputs_sheet(self, sheet, batch):
        """Parse the Inputs sheet.

        Strategy:
        1. Scan rows for period-level inputs (label in col A or B, value in next col).
        2. Detect the tabular GENCO section by looking for column headers.
        3. Parse tabular rows as GENCO operational data.
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        # Step 1: Find column headers for the operational table
        op_header_row_idx = None
        op_col_map = {}  # col_index -> field_name
        genco_col_idx = None

        for r_idx, row in enumerate(rows):
            row_labels = [str(c).lower().strip() if c is not None else '' for c in row]
            found_headers = {}
            found_genco_col = None
            for c_idx, label in enumerate(row_labels):
                for hdr_key, field_name in _OPERATIONAL_HEADER_MAP.items():
                    if hdr_key in label:
                        found_headers[c_idx] = field_name
                if 'genco' in label or 'plant' in label or 'station' in label:
                    found_genco_col = c_idx
            if len(found_headers) >= 2:
                op_header_row_idx = r_idx
                op_col_map = found_headers
                genco_col_idx = found_genco_col if found_genco_col is not None else 0
                break

        # Step 2: Parse period-level inputs (rows before operational table)
        search_limit = op_header_row_idx if op_header_row_idx else len(rows)
        for r_idx, row in enumerate(rows[:search_limit]):
            if not any(c for c in row):
                continue
            label_cell = str(row[0] or '').strip()
            if not label_cell:
                continue
            input_code = self._match_input_label(label_cell)
            if input_code:
                # Find first non-empty value after the label
                value = None
                for c_idx in range(1, len(row)):
                    if row[c_idx] is not None:
                        value = row[c_idx]
                        break
                if value is not None:
                    self.env['nbet.import.batch.line'].create({
                        'batch_id': batch.id,
                        'sheet_name': sheet.title,
                        'row_number': r_idx + 1,
                        'record_type': 'cycle_input',
                        'field_label': label_cell,
                        'raw_value': str(value),
                        'parsed_value_float': self._safe_float(value),
                        'input_type_code': input_code,
                        'status': 'mapped',
                    })

        # Step 3: Parse GENCO operational rows (after header row)
        if op_header_row_idx is not None:
            for r_idx in range(op_header_row_idx + 1, len(rows)):
                row = rows[r_idx]
                if not any(c for c in row):
                    continue
                genco_label = str(row[genco_col_idx] or '').strip() if genco_col_idx < len(row) else ''
                if not genco_label or genco_label.lower() in ('total', 'subtotal', ''):
                    continue
                for c_idx, field_name in op_col_map.items():
                    if c_idx >= len(row):
                        continue
                    cell_val = row[c_idx]
                    if cell_val is None:
                        continue
                    float_val = self._safe_float(cell_val)
                    # Unit conversion: if field is energy (kwh) and value looks like GWh
                    if 'kwh' in field_name and float_val < 100000 and float_val > 0:
                        # Likely GWh — convert to kWh
                        float_val *= 1_000_000
                    self.env['nbet.import.batch.line'].create({
                        'batch_id': batch.id,
                        'sheet_name': sheet.title,
                        'row_number': r_idx + 1,
                        'record_type': 'genco_data',
                        'field_label': field_name,
                        'raw_value': str(cell_val),
                        'parsed_value_float': float_val,
                        'participant_code': genco_label,
                        'status': 'mapped',
                    })

    def _parse_rates_sheet(self, sheet, batch):
        """Parse the Rates sheet for GENCO-specific rate data."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        # Find header row
        header_row_idx = None
        rate_col_map = {}
        genco_col_idx = 0

        for r_idx, row in enumerate(rows):
            row_labels = [str(c).lower().strip() if c is not None else '' for c in row]
            found = {}
            for c_idx, label in enumerate(row_labels):
                for hdr_key, field_name in _RATES_HEADER_MAP.items():
                    if hdr_key in label:
                        found[c_idx] = field_name
                if 'genco' in label or 'plant' in label or 'station' in label:
                    genco_col_idx = c_idx
            if found:
                header_row_idx = r_idx
                rate_col_map = found
                break

        if header_row_idx is None:
            self._log_error(batch, sheet.title, 0, '', 'mapping_error',
                            'Could not detect rate column headers in Rates sheet.')
            return

        for r_idx in range(header_row_idx + 1, len(rows)):
            row = rows[r_idx]
            if not any(c for c in row):
                continue
            genco_label = str(row[genco_col_idx] or '').strip() if genco_col_idx < len(row) else ''
            if not genco_label or genco_label.lower() in ('total', 'subtotal', ''):
                continue
            for c_idx, field_name in rate_col_map.items():
                if c_idx >= len(row):
                    continue
                cell_val = row[c_idx]
                if cell_val is None:
                    continue
                self.env['nbet.import.batch.line'].create({
                    'batch_id': batch.id,
                    'sheet_name': sheet.title,
                    'row_number': r_idx + 1,
                    'record_type': 'rate_data',
                    'field_label': field_name,
                    'raw_value': str(cell_val),
                    'parsed_value_float': self._safe_float(cell_val),
                    'participant_code': genco_label,
                    'status': 'mapped',
                })

    # ─────────────────────────────────────────────────────────────────────────
    # Import Committers
    # ─────────────────────────────────────────────────────────────────────────

    def _import_cycle_input_line(self, line, cycle):
        """Create or update a nbet.billing.cycle.input from a staged line."""
        input_type = self.env['nbet.billing.input.type'].search(
            [('code', '=', line.input_type_code)], limit=1
        )
        if not input_type:
            raise ValueError(f'Unknown input type code: {line.input_type_code}')
        existing = self.env['nbet.billing.cycle.input'].search([
            ('billing_cycle_id', '=', cycle.id),
            ('input_type_id', '=', input_type.id),
        ], limit=1)
        vals = {
            'billing_cycle_id': cycle.id,
            'input_type_id': input_type.id,
            'value_float': line.parsed_value_float,
            'source_sheet': line.sheet_name,
            'source_cell': f'Row {line.row_number}',
        }
        if existing:
            existing.write(vals)
        else:
            self.env['nbet.billing.cycle.input'].create(vals)

    def _import_genco_data_line(self, line, cycle):
        """Create or update nbet.genco.monthly.data from a staged line."""
        participant = self._match_participant(line.participant_code)
        if not participant:
            raise ValueError(f'Could not match participant: {line.participant_code}')
        existing = self.env['nbet.genco.monthly.data'].search([
            ('billing_cycle_id', '=', cycle.id),
            ('participant_id', '=', participant.id),
        ], limit=1)
        if not existing:
            existing = self.env['nbet.genco.monthly.data'].create({
                'billing_cycle_id': cycle.id,
                'participant_id': participant.id,
                'imported_from_file': True,
                'source_sheet': line.sheet_name,
                'source_row_no': line.row_number,
            })
        if line.field_label in dir(existing):
            existing.write({line.field_label: line.parsed_value_float})

    def _import_rate_data_line(self, line, cycle):
        """Create or update nbet.rate.snapshot from a staged rate line."""
        participant = self._match_participant(line.participant_code)
        if not participant:
            raise ValueError(f'Could not match participant: {line.participant_code}')
        existing = self.env['nbet.rate.snapshot'].search([
            ('billing_cycle_id', '=', cycle.id),
            ('participant_id', '=', participant.id),
            ('is_current', '=', True),
        ], limit=1)
        if not existing:
            existing = self.env['nbet.rate.snapshot'].create({
                'billing_cycle_id': cycle.id,
                'participant_id': participant.id,
            })
        if line.field_label and hasattr(existing, line.field_label):
            existing.write({line.field_label: line.parsed_value_float})

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _match_input_label(self, label):
        """Fuzzy-match a row label to an input_type_code."""
        normalized = label.lower().strip()
        for key, code in _INPUTS_LABEL_MAP.items():
            if key in normalized:
                return code
        return None

    def _match_participant(self, name_or_code):
        """Find a market participant by exact code or case-insensitive name."""
        if not name_or_code:
            return None
        # Try exact code match
        p = self.env['nbet.market.participant'].search(
            [('code', '=ilike', name_or_code.strip())], limit=1
        )
        if p:
            return p
        # Try name contains match
        p = self.env['nbet.market.participant'].search(
            [('name', 'ilike', name_or_code.strip())], limit=1
        )
        return p or None

    def _safe_float(self, value):
        """Safely convert a cell value to float."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(',', '').replace('%', '')
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    def _log_error(self, batch, sheet, row, cell, error_type, message, raw_value=''):
        self.env['nbet.import.error.log'].create({
            'batch_id': batch.id,
            'sheet_name': sheet,
            'row_number': row,
            'cell_ref': cell,
            'error_type': error_type,
            'message': message,
            'raw_value': raw_value,
        })

    def _build_preview_html(self, batch):
        """Build an HTML preview table of the staged lines."""
        lines = batch.line_ids[:50]  # preview first 50 rows
        rows_html = ''
        for line in lines:
            rows_html += (
                f'<tr>'
                f'<td>{line.sheet_name}</td>'
                f'<td>{line.row_number}</td>'
                f'<td>{line.record_type}</td>'
                f'<td>{line.participant_code or ""}</td>'
                f'<td>{line.field_label or ""}</td>'
                f'<td>{line.raw_value or ""}</td>'
                f'<td>{line.parsed_value_float:.4f}</td>'
                f'<td><span class="badge bg-{"success" if line.status == "mapped" else "warning"}">{line.status}</span></td>'
                f'</tr>'
            )
        error_count = batch.error_count
        error_note = f'<p class="text-danger"><strong>{error_count} error(s)</strong> during parsing.</p>' if error_count else ''
        return (
            f'{error_note}'
            f'<p>Showing first {len(lines)} of {batch.total_lines} staged rows.</p>'
            f'<table class="table table-sm table-bordered">'
            f'<thead><tr>'
            f'<th>Sheet</th><th>Row</th><th>Type</th><th>Participant</th>'
            f'<th>Field</th><th>Raw Value</th><th>Parsed Value</th><th>Status</th>'
            f'</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            f'</table>'
        )

    def action_reset(self):
        """Reset wizard to upload state, cancelling the current staged batch."""
        self.ensure_one()
        if self.batch_id and self.batch_id.state in ('draft', 'previewed'):
            self.batch_id.action_cancel()
        self.write({
            'state': 'draft',
            'batch_id': False,
            'excel_file': False,
            'excel_filename': False,
            'preview_html': False,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'nbet.excel.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
